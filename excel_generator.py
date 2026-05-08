import json
import logging
import re
import shutil
from pathlib import Path

import anthropic
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import ANTHROPIC_API_KEY, MODEL_FAST, PROMPTS_DIR, TEMPLATES_DIR

logger = logging.getLogger(__name__)
client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)

# Styles matching the user's template
HEADER_FONT = Font(name="Calibri", size=11, bold=True)
CELL_FONT = Font(name="Calibri", size=11)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CRITICAL_FONT = Font(name="Calibri", size=11, bold=True, color="C00000")


def _load_prompt() -> str:
    return (PROMPTS_DIR / "excel.txt").read_text(encoding="utf-8")


def _extract_json(text: str) -> dict:
    m = re.search(r"```(?:json)?\s*(\{.*?\})\s*```", text, re.DOTALL)
    if m:
        return json.loads(m.group(1))
    text = text.strip()
    if text.startswith("{"):
        return json.loads(text)
    start, end = text.find("{"), text.rfind("}")
    if start != -1 and end != -1:
        return json.loads(text[start:end + 1])
    raise ValueError("No se encontró JSON válido en la respuesta de Excel")


def _call_claude(analysis: dict) -> dict:
    prompt = (
        f"ANÁLISIS JSON:\n```json\n{json.dumps(analysis, ensure_ascii=False, indent=2)}\n```\n\n"
        + _load_prompt()
    )
    with client.messages.stream(
        model=MODEL_FAST,
        max_tokens=8192,
        messages=[{"role": "user", "content": prompt}],
    ) as stream:
        msg = stream.get_final_message()
    raw = "".join(block.text for block in msg.content if hasattr(block, "text"))
    return _extract_json(raw)


# ── Sheet writers ─────────────────────────────────────────────────────────────

def _style_header_row(ws, col_count: int) -> None:
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.alignment = WRAP
        cell.border = BORDER


def _fill_kv_sheet(ws, data: dict) -> None:
    """Fill a 2-column key/value sheet (rows already exist from template)."""
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        key_cell = row[0]
        if key_cell.value and str(key_cell.value) in data:
            val_cell = row[1]
            val_cell.value = data[str(key_cell.value)] or ""
            val_cell.font = CELL_FONT
            val_cell.alignment = WRAP
            val_cell.border = BORDER


def _append_rows_3col(ws, rows: list[dict], col_keys: list[str]) -> None:
    """Append dynamic rows to a 3-column sheet."""
    for row_data in rows:
        vals = [row_data.get(k, "") or "" for k in col_keys]
        ws.append(vals)
        row_idx = ws.max_row
        for col_i in range(1, 4):
            cell = ws.cell(row=row_idx, column=col_i)
            cell.font = CELL_FONT
            cell.alignment = WRAP
            cell.border = BORDER


def _append_rows_2col(ws, rows: list[dict], col_keys: list[str]) -> None:
    """Append dynamic rows to a 2-column sheet."""
    for row_data in rows:
        vals = [row_data.get(k, "") or "" for k in col_keys]
        ws.append(vals)
        row_idx = ws.max_row
        for col_i in range(1, 3):
            cell = ws.cell(row=row_idx, column=col_i)
            cell.font = CELL_FONT
            cell.alignment = WRAP
            cell.border = BORDER


def _set_col_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── Main generator ────────────────────────────────────────────────────────────

def generate_excel(analysis: dict, output_dir: str) -> str:
    logger.info("PASO 3: Extrayendo datos con Claude Sonnet para Excel…")
    data = _call_claude(analysis)
    logger.info("PASO 3: Datos recibidos. Construyendo Excel desde plantilla…")

    # Always use the user's template as base
    tpl = TEMPLATES_DIR / "excel" / "plantilla.xlsx"
    expediente = analysis.get("identificacion", {}).get("expediente", "licitacion")
    safe_exp = re.sub(r'[^\w\-]', '_', expediente)[:50]
    xlsx_path = str(Path(output_dir) / f"seguimiento_{safe_exp}.xlsx")

    if tpl.exists():
        shutil.copy2(str(tpl), xlsx_path)
        wb = openpyxl.load_workbook(xlsx_path)
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        # Create sheets from scratch with same names as template
        for name in ["Resumen licitación", "Cronograma", "Datos de contacto",
                     "Procedimiento", "Criterios", "Solvencias", "Penalidades",
                     "Condiciones especiales", "Notificaciones"]:
            wb.create_sheet(name)
        logger.warning("PASO 3: No se encontró plantilla.xlsx — Excel creado desde cero")

    sheet_map = {ws.title: ws for ws in wb.worksheets}

    # ── Resumen licitación ──
    if "Resumen licitación" in sheet_map:
        ws = sheet_map["Resumen licitación"]
        _fill_kv_sheet(ws, data.get("resumen", {}))
        _set_col_widths(ws, [42, 60])
        _style_header_row(ws, 2)

    # ── Cronograma ──
    if "Cronograma" in sheet_map:
        ws = sheet_map["Cronograma"]
        _fill_kv_sheet(ws, data.get("cronograma", {}))
        _set_col_widths(ws, [42, 30])
        _style_header_row(ws, 2)

    # ── Datos de contacto ──
    if "Datos de contacto" in sheet_map:
        ws = sheet_map["Datos de contacto"]
        _fill_kv_sheet(ws, data.get("contacto", {}))
        _set_col_widths(ws, [42, 60])
        _style_header_row(ws, 2)

    # ── Procedimiento ──
    if "Procedimiento" in sheet_map:
        ws = sheet_map["Procedimiento"]
        _fill_kv_sheet(ws, data.get("procedimiento", {}))
        _set_col_widths(ws, [42, 80])
        _style_header_row(ws, 2)

    # ── Criterios ──
    if "Criterios" in sheet_map:
        ws = sheet_map["Criterios"]
        _style_header_row(ws, 3)
        _append_rows_3col(ws, data.get("criterios", []),
                          ["Criterio de adjudicacion", "Ponderación máxima", "Regla / Fórmula"])
        _set_col_widths(ws, [50, 20, 70])

    # ── Solvencias ──
    if "Solvencias" in sheet_map:
        ws = sheet_map["Solvencias"]
        _style_header_row(ws, 3)
        _append_rows_3col(ws, data.get("solvencias", []),
                          ["Tipo", "Exigencia", "Observaciones"])
        _set_col_widths(ws, [25, 50, 60])

    # ── Penalidades ──
    if "Penalidades" in sheet_map:
        ws = sheet_map["Penalidades"]
        _style_header_row(ws, 3)
        _append_rows_3col(ws, data.get("penalidades", []),
                          ["Ámbito", "Supuesto", "Consecuencia/Importe"])
        _set_col_widths(ws, [25, 60, 40])

    # ── Condiciones especiales ──
    if "Condiciones especiales" in sheet_map:
        ws = sheet_map["Condiciones especiales"]
        _style_header_row(ws, 2)
        _append_rows_2col(ws, data.get("condiciones_especiales", []),
                          ["Ámbito", "Condición"])
        _set_col_widths(ws, [25, 100])

    # ── Notificaciones ──
    if "Notificaciones" in sheet_map:
        ws = sheet_map["Notificaciones"]
        _style_header_row(ws, 2)
        _append_rows_2col(ws, data.get("notificaciones", []),
                          ["Aspecto", "Detalle"])
        _set_col_widths(ws, [35, 80])

    # ── Sobres (1, 2 o 3 según el pliego) ──
    # Remove template's "Sobre único" and recreate dynamically
    sobres_data = data.get("sobres", [])
    tpl_sobre = "Sobre único"
    if tpl_sobre in sheet_map and sobres_data:
        del wb[tpl_sobre]

    for sobre in sobres_data:
        ws_name = sobre.get("nombre", "Sobre")[:31]
        ws = wb.create_sheet(title=ws_name)
        ws.append(["Bloque", "Documento / Requisito", "Detalle"])
        _style_header_row(ws, 3)
        _append_rows_3col(ws, sobre.get("filas", []),
                          ["Bloque", "Documento / Requisito", "Detalle"])
        _set_col_widths(ws, [25, 60, 80])

    # If no sobres data, keep original template sheet
    if not sobres_data and tpl_sobre in sheet_map:
        logger.warning("PASO 3: Sin datos de sobres, hoja 'Sobre único' permanece vacía")

    wb.save(xlsx_path)
    logger.info(f"PASO 3: Excel guardado en {xlsx_path}")
    return xlsx_path
